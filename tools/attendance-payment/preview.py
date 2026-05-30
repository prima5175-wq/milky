#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""설치 후 시트 모습을 보여주는 미리보기 xlsx 생성 (오늘=2026-05-30 가정)."""
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TODAY = datetime.date(2026, 5, 30)

C_DUR = {'60분': 'FCE4EC', '90분': 'D9EAD3', '120분': 'FFF2CC'}
C_USED = 'CFCFCF'; C_WEEK_OK = 'B6D7A8'; C_WEEK_MISS = 'EA9999'
C_COLHDR = '5B9BD5'; C_WEEKHDR = 'F9CB9C'; C_GRIDHDR = '9DC3E6'

# 열: A날짜 B이름 C학교학년 D전화 E등록여부 F금액 G등록회차 H형제할인 I등록일 J~V주차(13) W~회차
COL_SIB = 8; COL_REG = 9; WEEK0 = 10; WEEKN = 13; GRID0 = WEEK0 + WEEKN  # 23
GRID_COLS = 31
FREQ_PM = {'주1회': 4, '주2회': 8, '주3회': 12}

def fill(c): return PatternFill('solid', fgColor=c)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def plan_info(freq, dur, cycle):
    daily = freq == '매일반'
    per = GRID_COLS if daily else FREQ_PM[freq]
    rows = 3 if (daily or cycle == '분기') else 1
    return per, rows, daily

def week_counts(reg, dates, weeks):
    """주차별 출석수 list (등록일부터)."""
    out = []
    for i in range(weeks):
        ws = reg + datetime.timedelta(days=7*i)
        we = ws + datetime.timedelta(days=7)
        if ws > TODAY:
            out.append(None)  # 미래
        else:
            out.append(sum(1 for d in dates if ws <= d < we))
    return out

def build():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '명단(미리보기)'
    headers = ['테스트날짜','이름','학교/학년','휴대전화','등록여부','결제금액','등록회차','형제할인','등록일']
    for c,h in enumerate(headers,1):
        cell = ws.cell(1,c,h); cell.fill = fill(C_COLHDR); cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
    for i in range(WEEKN):
        cell = ws.cell(1, WEEK0+i, f'{i+1}주'); cell.fill = fill(C_WEEKHDR)
        cell.font = Font(color='783F04', bold=True, size=8); cell.alignment = Alignment(horizontal='center')
    for i in range(GRID_COLS):
        cell = ws.cell(1, GRID0+i, i+1); cell.fill = fill(C_GRIDHDR)
        cell.font = Font(color='1F4E79', bold=True, size=8); cell.alignment = Alignment(horizontal='center')

    # 샘플 학생: (날짜, 이름, 학교학년, 전화, 등록여부, 금액, freq, dur, cycle, 형제할인, 등록일, 출석날짜들)
    D = datetime.date
    students = [
        ('4/1','정채원','대도 4','010-7921','결제완료_정상등록',240000,'주1회','60분','월',False,
         D(2026,5,4),[D(2026,5,6),D(2026,5,13),D(2026,5,27)]),
        ('4/2','강서진','대도 6','010-8731','결제완료_정상등록',660000,'주2회','90분','분기',True,
         D(2026,3,10),[D(2026,3,11),D(2026,3,28),D(2026,4,8),D(2026,4,15),D(2026,4,22),
                       D(2026,5,6),D(2026,5,20),D(2026,5,27)]),
        ('4/6','김태윤','대도 4','010-7617','결제완료_정상등록',270000,'주2회','120분','월',False,
         D(2026,5,1),[D(2026,5,2),D(2026,5,8),D(2026,5,15),D(2026,5,22),D(2026,5,29)]),
        ('4/11','강지후','서울아카데미','010-4023','결제대기 중',990000,'매일반','90분','분기',False,
         D(2026,5,11),[D(2026,5,12),D(2026,5,13),D(2026,5,15),D(2026,5,19),D(2026,5,21),
                       D(2026,5,22),D(2026,5,26),D(2026,5,27),D(2026,5,28),D(2026,5,29)]),
    ]

    row = 2
    for st in students:
        (dt,name,sch,tel,reg,amt,freq,dur,cyc,sib,regd,dates) = st
        per, rows, daily = plan_info(freq,dur,cyc)
        price = round(amt*0.95) if sib else amt
        ws.cell(row,1,dt); ws.cell(row,2,name).font=Font(bold=True); ws.cell(row,3,sch); ws.cell(row,4,tel)
        rc = ws.cell(row,5,reg)
        rc.fill = fill({'결제완료_정상등록':'B6D7A8','결제대기 중':'FFE599','등록안함':'EA9999'}[reg])
        ws.cell(row,6,price).number_format='#,##0'
        # 등록회차 칩 느낌
        pc = ws.cell(row,7,f'{freq} {dur} {cyc if not daily else ""}'.strip())
        pc.fill = fill(C_DUR[dur]); pc.alignment=Alignment(horizontal='center'); pc.font=Font(bold=True,size=9)
        ws.cell(row,8,'☑' if sib else '☐').alignment=Alignment(horizontal='center')
        ws.cell(row,9,regd).number_format='yyyy-mm-dd'

        # 주차 띠
        weeks = 13 if (daily or cyc=='분기') else 5
        wc = week_counts(regd, dates, weeks)
        for i,cnt in enumerate(wc):
            cell = ws.cell(row, WEEK0+i)
            if cnt is None: continue
            cell.value = cnt; cell.alignment=Alignment(horizontal='center'); cell.font=Font(size=9)
            cell.fill = fill(C_WEEK_OK if cnt>0 else C_WEEK_MISS)

        # 회차 칸: per개씩 rows줄, 출석 날짜 채우고 회색
        seq = sorted(dates)
        di = 0
        for r in range(rows):
            n = GRID_COLS if daily else per
            for k in range(n):
                cell = ws.cell(row+r, GRID0+k)
                cell.fill = fill(C_DUR[dur]); cell.border = BORDER
                cell.alignment = Alignment(horizontal='center'); cell.font=Font(size=9)
            # 출석 날짜를 앞에서부터 채움(회색)
        # 날짜는 시간순으로 줄을 넘어가며 채움
        slots = []
        for r in range(rows):
            n = GRID_COLS if daily else per
            for k in range(n):
                slots.append((row+r, GRID0+k))
        for idx,d in enumerate(seq):
            if idx>=len(slots): break
            rr,cc = slots[idx]
            cell = ws.cell(rr,cc, d); cell.number_format='M/d'
            cell.fill = fill(C_USED); cell.alignment=Alignment(horizontal='center'); cell.font=Font(size=9)

        row += rows + 1  # 학생 사이 한 줄 띄움(미리보기 가독성)

    # 너비
    ws.column_dimensions['B'].width=10; ws.column_dimensions['C'].width=12
    ws.column_dimensions['D'].width=12; ws.column_dimensions['E'].width=15
    ws.column_dimensions['F'].width=10; ws.column_dimensions['G'].width=15
    ws.column_dimensions['H'].width=8; ws.column_dimensions['I'].width=11
    for i in range(WEEKN): ws.column_dimensions[get_column_letter(WEEK0+i)].width=4
    for i in range(GRID_COLS): ws.column_dimensions[get_column_letter(GRID0+i)].width=4.5
    ws.freeze_panes = 'C2'
    # 한 페이지 폭에 맞춰 인쇄(이미지 변환용)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.print_area = f'A1:{get_column_letter(GRID0+25)}{row}'

    # 범례 시트
    lg = wb.create_sheet('보는법')
    rows = [
        ('이 파일은 설치 후 시트 모습을 보여주는 미리보기입니다 (오늘=2026-05-30 가정).',None),
        ('',None),
        ('● 등록회차(G) 칩 색 = 시간', None),
        ('   60분 분홍 / 90분 초록 / 120분 노랑','C_DUR'),
        ('● 회차 칸(W열~) : 주N회×4칸, 분기납=3줄, 월납=1줄, 매일반=3줄 가득', None),
        ('   색칸=남은 회차, 회색칸=출석 완료(날짜 표시)', None),
        ('● 형제할인(H) ☑ = 결제금액 5% 할인 적용됨 (예: 강서진 660,000→627,000)', None),
        ('● 주차 띠(J~V) : 등록일부터 주 단위 출석 횟수', None),
        ('   초록=그 주 출석O, 빨강=그 주 결석(연락대상), 빈칸=아직 안 지난 주', None),
        ('', None),
        ('예) 정채원: 3주차가 빨강(0) → 그 주에 한 번도 안 옴', None),
        ('예) 강서진: 형제할인 적용 + 분기 13주 중 결석 주가 빨강으로 보임', None),
    ]
    for i,(t,_) in enumerate(rows):
        lg.cell(i+1,1,t)
    lg.cell(1,1).font=Font(bold=True,size=12); lg.column_dimensions['A'].width=80

    wb.save('미리보기.xlsx')
    print('saved 미리보기.xlsx')

build()
