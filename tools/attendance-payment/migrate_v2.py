#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
기존 학원 출결 엑셀 -> 새 [명단] 양식(Code.gs와 동일)으로 이관.

새 열 구조:
  A 테스트날짜 · B 이름 · C 학교/학년 · D 휴대전화 · E 등록여부 · F 결제금액
  G 등록회차 · H 형제할인 · I 등록일 · J~V 주차 띠(1~13) · W~ 회차 칸
  (54 _blk: 분기납 연속행 표시 CONT, 55 _orig, 56 원본참고, 57 검수)

사용:  python3 migrate_v2.py <기존.xlsx> <출력.xlsx>
가져온 뒤 구글시트에서 Code.gs 붙여넣고 [학원관리 ▸ 설치]를 1회 실행하면
드롭다운/체크박스/플랜단가가 적용됩니다.
"""
import sys, re, datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

YEAR = 2026
TODAY = datetime.date(2026, 5, 30)  # 주차 띠 기준(이관 시점)

# 새 레이아웃
A_DATE=1; B_NAME=2; C_SCH=3; D_TEL=4; E_REG=5; F_PRICE=6; G_PLAN=7
H_SIB=8; I_REG=9; WEEK0=10; WEEKN=5; GRID0=10+5  # 15 (주차 띠 한 줄=5주)
GRID_COLS=31
HELPER_COL=GRID0+GRID_COLS      # 46
HELPER_ORIG=HELPER_COL+1        # 47
MEMO_COL=HELPER_COL+2           # 48
FLAG_COL=HELPER_COL+3           # 49
CONT='CONT'

C_DUR={'60분':'FCE4EC','90분':'D9EAD3','120분':'FFF2CC'}
C_USED='CFCFCF'; C_OK='B6D7A8'; C_MISS='EA9999'; C_CONT='F3F3F3'
C_HDR='5B9BD5'; C_WHDR='F9CB9C'; C_GHDR='9DC3E6'
C_REG={'결제완료_정상등록':'B6D7A8','결제대기 중':'FFE599','등록안함':'EA9999'}
FREQ_PM={'주1회':4,'주2회':8,'주3회':12}

def fill(c): return PatternFill('solid', fgColor=c)
THIN=Side(style='thin', color='E0E0E0')
BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

# ---------- 파싱(기존 migrate.py와 동일 로직) ----------
def pick_plan_line(K):
    lines=[l.strip() for l in str(K).split('\n') if l.strip()]
    if not lines: return ''
    for l in reversed(lines):
        if ('회' in l) or ('몰독' in l) or ('매일' in l): return l
    return lines[-1]

def parse_plan(K):
    iss=dict(freq_missing=False,dur_default=False,cycle_missing=False,ambiguous=False)
    if not K or not str(K).strip():
        iss['freq_missing']=True; return (None,None,None,iss)
    raw=str(K); line=pick_plan_line(raw)
    iss['ambiguous']=('->' in raw) or ('변경' in raw) or ('?' in raw)
    if ('몰독' in line) or ('매일' in line): freq='매일반'
    elif '주1' in line: freq='주1회'
    elif '주2' in line: freq='주2회'
    elif '주3' in line: freq='주3회'
    else: freq=None; iss['freq_missing']=True
    if '120' in line: dur='120분'
    elif '90' in line: dur='90분'
    elif '60' in line: dur='60분'
    else: dur='90분'; iss['dur_default']=True
    if '분기' in line: cycle='분기'
    elif '월' in line: cycle='월'
    else: cycle=None; iss['cycle_missing']=True
    return (freq,dur,cycle,iss)

def parse_end_date(L):
    if isinstance(L,datetime.datetime): return L.date()
    if not L: return None
    for l in reversed([x.strip() for x in str(L).split('\n') if x.strip()]):
        m=re.search(r'(\d{1,2})\s*/\s*(\d{1,2})',l)
        if m:
            try: return datetime.date(YEAR,int(m.group(1)),int(m.group(2)))
            except ValueError: return None
    return None

def add_months(d,n):
    mo=d.month-1+n; y=d.year+mo//12; mo=mo%12+1
    dd=min(d.day,[31,29 if y%4==0 else 28,31,30,31,30,31,31,30,31,30,31][mo-1])
    return datetime.date(y,mo,dd)

def parse_remaining(O):
    if O is None: return (None,True)
    if isinstance(O,(int,float)): return (int(O),False)
    s=str(O)
    if '~' in s and not re.search(r'\d+\s*\(',s): return (None,True)
    nums=re.findall(r'\d+',s)
    if nums: return (int(nums[-1]), ('\n' in s or '(' in s))
    return (None,True)

def parse_amount(J):
    if J is None: return (None,False)
    if isinstance(J,(int,float)): return (int(J),False)
    s=str(J).replace(',',''); line=[l for l in s.split('\n') if l.strip()]
    line=line[0] if line else s
    m=re.search(r'(\d+)\s*만',line)
    if m: return (int(m.group(1))*10000,True)
    m=re.search(r'(\d{4,})',line)
    if m: return (int(m.group(1)), '\n' in s)
    return (None,True)

def read_students(path):
    wb=openpyxl.load_workbook(path,data_only=True); ws=wb['시트1']
    G0,GN=16,35
    name_rows=[r for r in range(3,ws.max_row+1) if ws.cell(r,2).value and str(ws.cell(r,2).value).strip()]
    name_rows.append(ws.max_row+1)
    out=[]
    for i in range(len(name_rows)-1):
        r,nxt=name_rows[i],name_rows[i+1]
        g=lambda c: ws.cell(r,c).value
        dates=[]
        for rr in range(r,nxt):
            for c in range(G0,GN+1):
                v=ws.cell(rr,c).value
                if isinstance(v,datetime.datetime): dates.append(v.date())
        dates.sort()
        out.append(dict(name=g(2),grade=g(3),school=g(4),phone=g(5),note=g(6),
            paydate=g(7),pay_jul=g(8),pay_jun=g(9),amount=g(10),plan=g(11),
            period=g(12),method=g(13),process=g(14),remain=g(15),
            portfolio=ws.cell(r,36).value,consult=ws.cell(r,37).value,dates=dates))
    return out

def week_counts(reg,dates,weeks):
    out=[]
    for i in range(weeks):
        ws=reg+datetime.timedelta(days=7*i); we=ws+datetime.timedelta(days=7)
        out.append(None if ws>TODAY else sum(1 for d in dates if ws<=d<we))
    return out

# ---------- 쓰기 ----------
def build(students,out):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title='명단'
    # 헤더
    head=['테스트날짜','이름','학교/학년','휴대전화','등록여부','결제금액','등록회차','형제할인','등록일']
    for c,h in enumerate(head,1):
        cell=ws.cell(1,c,h); cell.fill=fill(C_HDR); cell.font=Font(color='FFFFFF',bold=True)
        cell.alignment=Alignment(horizontal='center',vertical='center')
    for i in range(WEEKN):
        cell=ws.cell(1,WEEK0+i,f'{i+1}주'); cell.fill=fill(C_WHDR)
        cell.font=Font(color='783F04',bold=True,size=8); cell.alignment=Alignment(horizontal='center')
    for i in range(GRID_COLS):
        cell=ws.cell(1,GRID0+i,i+1); cell.fill=fill(C_GHDR)
        cell.font=Font(color='1F4E79',bold=True,size=8); cell.alignment=Alignment(horizontal='center')
    ws.cell(1,HELPER_COL,'_blk'); ws.cell(1,HELPER_ORIG,'_orig')
    ws.cell(1,MEMO_COL,'원본(참고)').font=Font(bold=True); ws.cell(1,FLAG_COL,'검수').font=Font(bold=True)

    flagged=0; row=2
    for st in students:
        freq,dur,cycle,iss=parse_plan(st['plan'])
        due=parse_end_date(st['period'])
        rem,rflag=parse_remaining(st['remain'])
        amt,aflag=parse_amount(st['amount'])
        session=freq not in (None,'매일반')
        flag=(iss['freq_missing'] or iss['ambiguous'] or (session and rem is None)
              or (st['amount'] not in (None,'') and amt is None))

        # 등록회차 문자열(드롭다운 항목과 일치)
        if freq=='매일반':
            cyc='분기'; gval=f'매일반 {dur}'
        elif freq:
            cyc = '분기' if (cycle=='분기' or cycle is None) else '월'
            gval=f'{freq} {dur} {cyc}'
        else:
            cyc=None; gval=''

        daily=(freq=='매일반')
        months = 3 if (daily or cyc=='분기') else 1
        rows = 3 if (daily or cyc=='분기') else 1
        per = GRID_COLS if daily else (FREQ_PM.get(freq,0))

        # 등록일 = 기간종료 - 개월(현재 등록 기간 시작)
        regd = add_months(due,-months) if (due and freq) else None
        # 현재 기간 출석만
        dates=st['dates']
        if regd:
            cur=[d for d in dates if d>=regd]
            dates = cur if cur else (cur if rem is not None else dates)

        # 등록여부
        proc=str(st['process'] or '')
        if not freq: regstat='등록안함'
        elif ('발송' in proc) or ('대기' in proc) or ('문자' in proc): regstat='결제대기 중'
        else: regstat='결제완료_정상등록'

        # ---- 한 줄(소유행) 기본정보 ----
        ws.cell(row,B_NAME,st['name']).font=Font(bold=True)
        sch=' '.join(str(x) for x in [st['school'],int(st['grade']) if isinstance(st['grade'],(int,float)) else st['grade']] if x not in (None,''))
        ws.cell(row,C_SCH,sch)
        ws.cell(row,D_TEL,st['phone'])
        rc=ws.cell(row,E_REG,regstat); rc.fill=fill(C_REG[regstat]); rc.alignment=Alignment(horizontal='center')
        if amt is not None: ws.cell(row,F_PRICE,amt).number_format='#,##0'
        if gval:
            pc=ws.cell(row,G_PLAN,gval); pc.fill=fill(C_DUR.get(dur,C_DUR['90분']))
            pc.alignment=Alignment(horizontal='center'); pc.font=Font(bold=True,size=9)
        ws.cell(row,H_SIB,False)
        if regd: ws.cell(row,I_REG,regd).number_format='yyyy-mm-dd'

        # 원본 보존 + 검수
        memo=[]
        memo.append('플랜:'+(str(st['plan']).replace('\n',' / ') if st['plan'] else '없음'))
        if st['period']: memo.append('기간:'+str(st['period']).replace('\n',' / '))
        if st['remain'] not in (None,''): memo.append('남은:'+str(st['remain']).replace('\n',' / '))
        if st['amount']: memo.append('금액:'+str(st['amount']).replace('\n',' / '))
        pay=[str(x).replace('\n',' ') for x in [st['paydate'],st['pay_jul'],st['pay_jun'],st['process']] if x]
        if pay: memo.append('결제:'+' | '.join(pay))
        extra=[str(x).replace('\n',' ') for x in [st['note'],st['consult'],st['portfolio']] if x]
        if extra: memo.append('메모:'+' | '.join(extra))
        ws.cell(row,MEMO_COL,' || '.join(memo))
        if flag:
            tags=[]
            if iss['freq_missing']: tags.append('회차없음')
            if iss['ambiguous']: tags.append('변경표기')
            if session and rem is None: tags.append('남은회차?')
            if st['amount'] not in (None,'') and amt is None: tags.append('금액?')
            ws.cell(row,FLAG_COL,'⚠️ '+','.join(tags)); flagged+=1

        # 연속행(분기/매일반) CONT 표시
        for r in range(1,rows):
            rr=row+r
            ws.cell(rr,HELPER_COL,CONT)
            for c in range(1,I_REG+1): ws.cell(rr,c).fill=fill(C_CONT)

        # ---- 주차 띠 (한 줄=한 달, 5주씩) ----
        if regd and freq:
            for w,cnt in enumerate(week_counts(regd,dates,WEEKN*rows)):
                if cnt is None: continue
                rr=row+w//WEEKN; cc=WEEK0+w%WEEKN
                cell=ws.cell(rr,cc,cnt); cell.alignment=Alignment(horizontal='center')
                cell.font=Font(size=9); cell.fill=fill(C_OK if cnt>0 else C_MISS)

        # ---- 회차 칸 ----
        if freq:
            durc=C_DUR.get(dur,C_DUR['90분'])
            slots=[]
            for r in range(rows):
                n=GRID_COLS if daily else per
                for k in range(n):
                    cell=ws.cell(row+r,GRID0+k); cell.fill=fill(durc); cell.border=BORDER
                    cell.alignment=Alignment(horizontal='center'); cell.font=Font(size=9)
                    slots.append((row+r,GRID0+k))
            for idx,d in enumerate(sorted(dates)):
                if idx>=len(slots): break
                rr,cc=slots[idx]
                cell=ws.cell(rr,cc,d); cell.number_format='M/d'; cell.fill=fill(C_USED)

        row += rows

    # 너비/숨김
    for col,w in [('B',10),('C',12),('D',13),('E',15),('F',10),('G',14),('H',8),('I',11)]:
        ws.column_dimensions[col].width=w
    for i in range(WEEKN): ws.column_dimensions[get_column_letter(WEEK0+i)].width=4
    for i in range(GRID_COLS): ws.column_dimensions[get_column_letter(GRID0+i)].width=4.5
    ws.column_dimensions[get_column_letter(HELPER_COL)].hidden=True
    ws.column_dimensions[get_column_letter(HELPER_ORIG)].hidden=True
    ws.column_dimensions[get_column_letter(MEMO_COL)].width=60
    ws.freeze_panes='C2'

    wb.save(out)
    return flagged

if __name__=='__main__':
    src,out=sys.argv[1],sys.argv[2]
    sts=read_students(src); n=build(sts,out)
    print('학생 %d명 이관(새 구조), ⚠️검수 %d명 -> %s'%(len(sts),n,out))
