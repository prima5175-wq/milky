#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
기존 학원 출결 엑셀 -> 새 [출결관리] 양식으로 자동 이관.

사용:  python3 migrate.py <기존.xlsx> <출력.xlsx>

특징
 - 신원(이름/학년/학교/연락처)·출석 날짜는 그대로 옮김.
 - K(회차)/L(기간)/O(남은회차)/금액 등은 best-effort 파싱.
 - 원본 값은 손실 없이 [결제메모] 칸에 그대로 보존하고, 모호하면 ⚠️검수 표시.
 - 시간별 색칠(60분 분홍/90분 초록/120분 노랑), 출석=회색.
 - [플랜단가][대시보드][사용안내] 시트도 함께 생성 → 구글시트로 바로 import 가능.
"""
import sys, re, datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

YEAR = 2026  # 원본 데이터 기준 연도

# 새 레이아웃 (Code.gs와 동일)
COL = dict(NUM=1, NAME=2, GRADE=3, SCHOOL=4, PHONE=5, NOTE=6, REGDATE=7,
           FREQ=8, DUR=9, CYCLE=10, PRICE=11, TOTAL=12, LEFT=13, DUE=14,
           PAYMETHOD=15, PAYMEMO=16, WPW=17)
GRID_START = 18
GRID_COLS = 31
ROWS_PER = 3
HDR = 2
DATA0 = 3

C_DUR = {'60분': 'FCE4EC', '90분': 'D9EAD3', '120분': 'FFF2CC'}
C_USED = 'CFCFCF'; C_DUE = 'FDE9D9'
C_HEADER = '434343'; C_COLHDR = '5B9BD5'; C_GRIDHDR = '9DC3E6'

FREQ_WPW = {'주1회': 1, '주2회': 2, '주3회': 3, '매일반': 0}

# ---------- 파싱 ----------
def pick_plan_line(K):
    lines = [l.strip() for l in str(K).split('\n') if l.strip()]
    if not lines:
        return ''
    # '회'나 '몰독'이 든 마지막 줄(=최근 등록)을 우선
    for l in reversed(lines):
        if ('회' in l) or ('몰독' in l) or ('매일' in l):
            return l
    return lines[-1]

def parse_plan(K):
    """returns (freq, dur, cycle, issues) where issues is a dict of soft notes/flags"""
    issues = dict(freq_missing=False, dur_default=False, cycle_missing=False, ambiguous=False)
    if not K or not str(K).strip():
        issues['freq_missing'] = True
        return (None, None, None, issues)
    raw = str(K)
    line = pick_plan_line(raw)
    issues['ambiguous'] = ('->' in raw) or ('변경' in raw) or ('?' in raw)
    # 횟수
    if ('몰독' in line) or ('매일' in line):
        freq = '매일반'
    elif '주1' in line:
        freq = '주1회'
    elif '주2' in line:
        freq = '주2회'
    elif '주3' in line:
        freq = '주3회'
    else:
        freq = None; issues['freq_missing'] = True
    # 시간 (미표기 → 90분 기본값)
    if '120' in line:
        dur = '120분'
    elif '90' in line:
        dur = '90분'
    elif '60' in line:
        dur = '60분'
    else:
        dur = '90분'; issues['dur_default'] = True
    # 납부 (미표기 → 비워둠, 원장이 선택)
    if '분기' in line:
        cycle = '분기납'
    elif '월' in line:
        cycle = '월납'
    else:
        cycle = None; issues['cycle_missing'] = True
    return (freq, dur, cycle, issues)

def parse_end_date(L):
    if isinstance(L, datetime.datetime):
        return L
    if not L:
        return None
    lines = [l.strip() for l in str(L).split('\n') if l.strip()]
    for l in reversed(lines):
        m = re.search(r'(\d{1,2})\s*/\s*(\d{1,2})', l)
        if m:
            mo, d = int(m.group(1)), int(m.group(2))
            try:
                return datetime.datetime(YEAR, mo, d)
            except ValueError:
                return None
    return None

def add_months(date, n):
    mo = date.month - 1 + n
    y = date.year + mo // 12
    mo = mo % 12 + 1
    d = min(date.day, [31,29 if y%4==0 else 28,31,30,31,30,31,31,30,31,30,31][mo-1])
    return datetime.datetime(y, mo, d)

def parse_remaining(O):
    """returns (int or None, flag)"""
    if O is None:
        return (None, True)
    if isinstance(O, (int, float)):
        return (int(O), False)
    s = str(O)
    if '~' in s and not re.search(r'\d+\s*\(', s):  # 날짜형(몰독) → 회차 아님
        return (None, True)
    nums = re.findall(r'\d+', s)
    if nums:
        return (int(nums[-1]), '\n' in s or '(' in s)
    return (None, True)

def parse_amount(J):
    if J is None:
        return (None, False)
    if isinstance(J, (int, float)):
        return (int(J), False)
    s = str(J).replace(',', '')
    line = [l for l in s.split('\n') if l.strip()]
    line = line[0] if line else s
    m = re.search(r'(\d+)\s*만', line)
    if m:
        return (int(m.group(1)) * 10000, True)
    m = re.search(r'(\d{4,})', line)
    if m:
        return (int(m.group(1)), '\n' in s)
    return (None, True)

# ---------- 읽기 ----------
def read_students(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['시트1']
    G0, GN = 16, 35  # 원본 출석 그리드 P..AI
    name_rows = [r for r in range(3, ws.max_row+1)
                 if ws.cell(r,2).value and str(ws.cell(r,2).value).strip()]
    name_rows.append(ws.max_row+1)
    students = []
    for i in range(len(name_rows)-1):
        r, nxt = name_rows[i], name_rows[i+1]
        g = lambda c: ws.cell(r, c).value
        dates = []
        for rr in range(r, nxt):
            for c in range(G0, GN+1):
                v = ws.cell(rr, c).value
                if isinstance(v, datetime.datetime):
                    dates.append(v)
        dates.sort()
        students.append(dict(
            name=g(2), grade=g(3), school=g(4), phone=g(5), note=g(6),
            paydate=g(7), pay_jul=g(8), pay_jun=g(9), amount=g(10),
            plan=g(11), period=g(12), method=g(13), process=g(14), remain=g(15),
            portfolio=ws.cell(r,36).value, consult=ws.cell(r,37).value,
            dates=dates,
        ))
    return students

# ---------- 쓰기 ----------
def fill(color):
    return PatternFill('solid', fgColor=color)

THIN = Side(style='thin', color='E0E0E0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def build(students, out):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '출결관리'
    last_col = GRID_START + GRID_COLS - 1

    # 제목
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    t = ws.cell(1, 1, '출결·결제 관리   |   60분 분홍·90분 초록·120분 노랑   ·   날짜=출석(회색)   ·   남은 색칸=남은 회차   ·   ⚠️결제메모=원본 보존/검수')
    t.fill = fill(C_HEADER); t.font = Font(color='FFFFFF', bold=True)

    headers = ['번호','이름','학년','학교','연락처','특이사항','등록일','횟수','시간',
               '납부','금액','총회차','남은회차','다음결제일','결제수단','결제메모','주당']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(HDR, c, h)
        cell.fill = fill(C_COLHDR); cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for i in range(GRID_COLS):
        cell = ws.cell(HDR, GRID_START+i, i+1)
        cell.fill = fill(C_GRIDHDR); cell.font = Font(color='1F4E79', bold=True, size=8)
        cell.alignment = Alignment(horizontal='center')

    flagged = 0
    for idx, st in enumerate(students):
        top = DATA0 + idx*ROWS_PER
        freq, dur, cycle, issues = parse_plan(st['plan'])
        due = parse_end_date(st['period'])
        rem, rflag = parse_remaining(st['remain'])
        amt, aflag = parse_amount(st['amount'])
        # 진짜 검수 필요한 경우만 ⚠️ : 횟수 못 읽음 / 세션제 인데 남은회차 숫자 없음 / 금액 파싱 실패 / 변경표기
        session_based = freq not in (None, '매일반')
        flag = (issues['freq_missing'] or issues['ambiguous']
                or (session_based and rem is None)
                or (st['amount'] not in (None, '') and amt is None))
        soft = []
        if issues['dur_default']: soft.append('시간기본90분')
        if issues['cycle_missing']: soft.append('납부확인')

        ws.cell(top, COL['NUM'], idx+1)
        ws.cell(top, COL['NAME'], st['name'])
        ws.cell(top, COL['GRADE'], st['grade'])
        ws.cell(top, COL['SCHOOL'], st['school'])
        ws.cell(top, COL['PHONE'], st['phone'])
        # 특이사항 = 원본 특이사항 + 상담 + 포폴
        note_parts = [str(x) for x in [st['note'], st['consult'], st['portfolio']] if x]
        if note_parts:
            ws.cell(top, COL['NOTE'], ' / '.join(note_parts))
        if freq: ws.cell(top, COL['FREQ'], freq)
        if dur:  ws.cell(top, COL['DUR'], dur)
        if cycle: ws.cell(top, COL['CYCLE'], cycle)
        if amt is not None:
            ws.cell(top, COL['PRICE'], amt).number_format = '#,##0'
        ws.cell(top, COL['WPW'], FREQ_WPW.get(freq, 0))
        if due:
            dc = ws.cell(top, COL['DUE'], due); dc.number_format = 'yyyy-mm-dd'
        ws.cell(top, COL['PAYMETHOD'], st['method'])

        # 결제메모(원본 보존)
        memo = []
        if flag: memo.append('⚠️검수')
        if soft: memo.append('(' + ','.join(soft) + ')')
        memo.append('원본플랜:' + (str(st['plan']).replace('\n',' / ') if st['plan'] else '없음'))
        if st['period']: memo.append('기간:' + str(st['period']).replace('\n',' / '))
        if st['remain'] not in (None,''): memo.append('남은:' + str(st['remain']).replace('\n',' / '))
        if st['amount']: memo.append('금액:' + str(st['amount']).replace('\n',' / '))
        pay = [str(x).replace('\n',' ') for x in [st['paydate'], st['pay_jul'], st['pay_jun'], st['process']] if x]
        if pay: memo.append('결제이력:' + ' | '.join(pay))
        ws.cell(top, COL['PAYMEMO'], '\n'.join(memo)).alignment = Alignment(wrap_text=False, vertical='top')

        # ---- 회차 칸 채우기 ----
        # 새 양식은 '현재 등록 기간' 기준 → 결제일(기간 종료) - 개월 이후의 출석만 표시
        months = 3 if (cycle == '분기납' or freq == '매일반' or cycle is None) else 1
        dates = st['dates']
        if due is not None:
            start = add_months(due, -months)
            cur = [d for d in dates if d >= start]
            # 현재 기간 출석이 하나도 없고 과거 기록만 있으면(예: 미등록 기간) 전체 유지
            dates = cur if cur else (dates if rem is None else cur)
        used = len(dates)
        capacity = ROWS_PER * GRID_COLS
        if freq == '매일반':
            ws.cell(top, COL['TOTAL'], '매일반')
            ws.cell(top, COL['LEFT'], '출석 %d' % used)
            seq = [('date', d) for d in dates[:capacity]]
        else:
            total = used + (rem if rem is not None else 0)
            ws.cell(top, COL['TOTAL'], total if total else (used or ''))
            ws.cell(top, COL['LEFT'], rem if rem is not None else '')
            seq = [('date', d) for d in dates]
            if rem and rem > 0:
                seq += [('slot', None)] * rem
            seq = seq[:capacity]

        durcolor = C_DUR.get(dur, C_DUR['90분'])
        for k, (kind, val) in enumerate(seq):
            rr = top + k // GRID_COLS
            cc = GRID_START + k % GRID_COLS
            cell = ws.cell(rr, cc)
            if kind == 'date':
                cell.value = val; cell.number_format = 'M/d'
                cell.fill = fill(C_USED)
            else:
                cell.fill = fill(durcolor)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(size=9)

        if flag:
            flagged += 1

    # 데이터 검증(드롭다운)
    nrows = len(students) * ROWS_PER
    end = DATA0 + nrows - 1
    def add_dv(col, items):
        dv = DataValidation(type='list', formula1='"%s"' % ','.join(items), allow_blank=True)
        ws.add_data_validation(dv)
        L = get_column_letter(col)
        dv.add('%s%d:%s%d' % (L, DATA0, L, end))
    add_dv(COL['FREQ'], ['주1회','주2회','주3회','매일반'])
    add_dv(COL['DUR'], ['60분','90분','120분'])
    add_dv(COL['CYCLE'], ['월납','분기납'])

    # 조건부 서식: 다음결제일 임박/지남
    N = get_column_letter(COL['DUE'])
    rng = '%s%d:%s%d' % (N, DATA0, N, end)
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['AND($%s%d<>"",$%s%d<=TODAY())' % (N, DATA0, N, DATA0)],
        fill=fill('F4CCCC')))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['AND($%s%d<>"",$%s%d<=TODAY()+7)' % (N, DATA0, N, DATA0)],
        fill=fill('FCE5CD')))

    # 보기 설정
    ws.freeze_panes = 'C3'
    ws.column_dimensions[get_column_letter(COL['WPW'])].hidden = True
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions[get_column_letter(COL['PAYMEMO'])].width = 40
    for c in range(GRID_START, last_col+1):
        ws.column_dimensions[get_column_letter(c)].width = 5

    build_price(wb)
    build_dashboard(wb)
    build_help(wb)
    wb.save(out)
    return flagged

def build_price(wb):
    sh = wb.create_sheet('플랜단가')
    sh.append(['횟수','시간','납부','금액(편집하세요)'])
    for c in range(1,5):
        sh.cell(1,c).fill = fill(C_HEADER); sh.cell(1,c).font = Font(color='FFFFFF', bold=True)
    base = [
        ('주1회','60분',120000,342000),('주1회','90분',160000,480000),('주1회','120분',210000,630000),
        ('주2회','60분',180000,513000),('주2회','90분',220000,660000),('주2회','120분',270000,810000),
        ('주3회','60분',240000,684000),('주3회','90분',300000,855000),('주3회','120분',360000,999000),
        ('매일반','60분',300000,855000),('매일반','90분',350000,990000),('매일반','120분',420000,1190000),
    ]
    for f,d,m,q in base:
        sh.append([f,d,'월납',m]); sh.append([f,d,'분기납',q])
    for r in range(2, sh.max_row+1):
        sh.cell(r,4).number_format = '#,##0'
    sh.freeze_panes = 'A2'
    sh.cell(1,6,'※ 횟수+시간+납부 조합으로 금액 자동 조회. 실제 금액으로 수정하세요.').font = Font(color='888888')

def build_dashboard(wb):
    sh = wb.create_sheet('대시보드')
    M = "'출결관리'"
    sh.cell(1,1,'📊 운영 대시보드 (자동 갱신)').font = Font(size=14, bold=True)
    labels = ['총 등록생','이번달 결제예정 합계','결제 임박(7일내) 인원','회차 소진 위험 인원']
    for i,l in enumerate(labels):
        sh.cell(3+i,1,l).font = Font(bold=True)
    sh.cell(3,2).value = f'=COUNTIF({M}!B3:B,"?*")'
    sh.cell(4,2).value = (f'=SUMPRODUCT(({M}!N3:N>=DATE(YEAR(TODAY()),MONTH(TODAY()),1))*'
                          f'({M}!N3:N<=EOMONTH(TODAY(),0))*N({M}!K3:K))')
    sh.cell(4,2).number_format = '#,##0'
    sh.cell(5,2).value = f'=SUMPRODUCT(({M}!B3:B<>"")*({M}!N3:N<>"")*({M}!N3:N<=TODAY()+7))'
    sh.cell(6,2).value = (f'=SUMPRODUCT(({M}!B3:B<>"")*({M}!Q3:Q>0)*(N({M}!M3:M)>0)*'
                          f'((({M}!N3:N-TODAY())/7)*{M}!Q3:Q<N({M}!M3:M)))')
    sh.cell(8,1,'💳 결제 임박 / 지난 학생').font = Font(bold=True)
    for i,h in enumerate(['이름','연락처','다음결제일','금액','D-day']):
        sh.cell(9,1+i,h).font = Font(bold=True)
    sh.cell(10,1).value = (f'=IFERROR(SORT(FILTER({{{M}!B3:B,{M}!E3:E,{M}!N3:N,{M}!K3:K,{M}!N3:N-TODAY()}},'
                           f'({M}!B3:B<>"")*({M}!N3:N<>"")*({M}!N3:N<=TODAY()+7)),3,TRUE),"결제 임박 학생 없음")')
    sh.cell(8,7,'⚠️ 회차 소진 위험').font = Font(bold=True)
    for i,h in enumerate(['이름','연락처','남은회차','다음결제일','남은일']):
        sh.cell(9,7+i,h).font = Font(bold=True)
    sh.cell(10,7).value = (f'=IFERROR(SORT(FILTER({{{M}!B3:B,{M}!E3:E,{M}!M3:M,{M}!N3:N,{M}!N3:N-TODAY()}},'
                           f'({M}!B3:B<>"")*({M}!Q3:Q>0)*(N({M}!M3:M)>0)*'
                           f'((({M}!N3:N-TODAY())/7)*{M}!Q3:Q<N({M}!M3:M))),5,TRUE),"소진 위험 학생 없음")')

def build_help(wb):
    sh = wb.create_sheet('사용안내')
    lines = [
        '📚 사용 안내 (이관본)',
        '',
        '이 파일은 기존 엑셀을 새 양식으로 자동 변환한 결과입니다.',
        '',
        '[중요] 결제메모 칸에 원본 데이터를 그대로 보존했습니다.',
        '  · ⚠️검수 표시가 있는 학생은 자동 변환이 모호했던 경우이니 횟수/시간/납부/남은회차를 확인하세요.',
        '  · 시간 미표기 플랜은 90분으로 기본 설정했습니다.',
        '',
        '[색상] 60분 분홍 · 90분 초록 · 120분 노랑 · 회색=출석 완료 · 남은 색칸=남은 회차',
        '',
        '[자동화 켜기] 구글시트로 import 후, 확장프로그램 ▸ Apps Script 에 Code.gs 붙여넣기.',
        '  · 이후 횟수/시간/납부 드롭다운을 바꾸면 회차 칸이 새로 생성됩니다(기존 출석 초기화 주의).',
        '  · 출석 체크: 칸에 날짜 입력 또는 메뉴 [학원관리 ▸ 오늘 출석 체크].',
        '',
        '[다음결제일] 기존 기간(L열) 종료일을 그대로 옮겼습니다. 7일 이내/지난 학생은 빨강 강조됩니다.',
    ]
    for i,l in enumerate(lines):
        sh.cell(i+1,1,l)
    sh.cell(1,1).font = Font(size=14, bold=True)
    sh.column_dimensions['A'].width = 90

if __name__ == '__main__':
    src, out = sys.argv[1], sys.argv[2]
    sts = read_students(src)
    n = build(sts, out)
    print('학생 %d명 이관 완료, ⚠️검수 필요 %d명 -> %s' % (len(sts), n, out))
